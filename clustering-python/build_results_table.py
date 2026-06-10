"""
build_results_table.py
======================
Consolida los CSVs `pred_<ALGO>.csv` de la carpeta `predictions/` en una
tabla Excel + CSV listas para usar en tu tesis.

Uso básico (cuando solo tienes los resultados de CLIP en predictions/):
    python build_results_table.py --model CLIP

Cuando quieras agregar otros modelos, organiza así:
    predictions/
        clip/        <- mueve aquí los pred_*.csv que generaste con CLIP
            pred_BAT.csv
            pred_CSCLP.csv
            ...
        blip/        <- los nuevos cuando los generes
            pred_BAT.csv
            ...
        qwen_vl/
            ...
        internvl/
            ...

Y luego corre:
    python build_results_table.py --model CLIP
    python build_results_table.py --model BLIP
    python build_results_table.py --all       # combina TODOS los modelos

Genera:
    results_<modelo>.xlsx       (5 hojas: Resumen + ARI + AMI + NMI + Silhouette + Datos_completos)
    results_<modelo>_long.csv   (formato largo para análisis)
    results_all_models.xlsx     (cuando se usa --all, compara modelos)
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Mapping de algoritmos: nombre_visible -> nombre_archivo
# ----------------------------------------------------------------------------
ALGO_FILES = {
    'BAT':       'pred_BAT.csv',
    'CSCLP':     'pred_CSCLP.csv',
    'K-MEDOIDS': 'pred_KMEDOIDS.csv',
    'ACO':       'pred_ACO.csv',
    'PSO':       'pred_PSO.csv',
    'HCAKC':     'pred_HCAKC.csv',
    'KM-MILP':   'pred_KM-MILP.csv',
    'SCK1':      'pred_SCK1.csv',
}

PROJECT_ROOT    = Path(__file__).resolve().parent
PREDICTIONS_DIR = PROJECT_ROOT / 'predictions'


# ============================================================================
# Estilos Excel
# ============================================================================

HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='2F5597')
CELL_FONT   = Font(name='Arial', size=10)
BEST_FILL   = PatternFill('solid', start_color='C6EFCE')
TITLE_FONT  = Font(name='Arial', bold=True, size=14)
BOLD_FONT   = Font(name='Arial', bold=True, size=10)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)


# ============================================================================
# Lectura de datos
# ============================================================================

def find_predictions_dir(model: str) -> Path:
    """
    Resuelve el directorio de predictions para un modelo dado.
    Intenta primero `predictions/<model>/`, luego `predictions/` (estructura plana).
    """
    nested = PREDICTIONS_DIR / model.lower()
    if nested.exists() and any(nested.glob('pred_*.csv')):
        return nested
    if PREDICTIONS_DIR.exists() and any(PREDICTIONS_DIR.glob('pred_*.csv')):
        print(f'⚠ Usando estructura plana en {PREDICTIONS_DIR}. Sugiero '
              f'reorganizar a {PREDICTIONS_DIR}/{model.lower()}/ cuando '
              f'tengas más modelos.')
        return PREDICTIONS_DIR
    raise FileNotFoundError(
        f'No encontré CSVs en {nested} ni en {PREDICTIONS_DIR}. '
        f'Verifica la ruta.'
    )


def load_model_results(model: str) -> pd.DataFrame:
    """Lee los CSVs de un modelo y devuelve un DataFrame en formato largo."""
    pred_dir = find_predictions_dir(model)
    rows: list[dict] = []

    for algo_name, fname in ALGO_FILES.items():
        csv_path = pred_dir / fname
        if not csv_path.exists():
            print(f'  ⚠ Falta {csv_path.name} (saltando {algo_name})')
            continue
        df = pd.read_csv(csv_path)
        for _, row in df.iterrows():
            rows.append({
                'Modelo':     model,
                'Dataset':    row['name'],
                'Algoritmo':  algo_name,
                'ARI':        round(row['ARI'], 4),
                'AMI':        round(row['AMI'], 4),
                'NMI':        round(row['NMI'], 4),
                'Silhouette': round(row['Silhouette_mean'], 4),
            })

    if not rows:
        raise RuntimeError(f'No se encontraron resultados para modelo {model}')

    return pd.DataFrame(rows)


# ============================================================================
# Escritura de Excel
# ============================================================================

def write_metric_sheet(
    wb: Workbook, metric: str, df_long: pd.DataFrame, model: str
) -> None:
    """Crea una hoja Excel con datasets en filas y algoritmos en columnas."""
    ws = wb.create_sheet(metric)

    datasets = sorted(df_long['Dataset'].unique())
    algos    = [a for a in ALGO_FILES if a in df_long['Algoritmo'].unique()]

    pivot = df_long.pivot(index='Dataset', columns='Algoritmo', values=metric)
    pivot = pivot.reindex(index=datasets, columns=algos)

    # Título
    ws.cell(row=1, column=1, value=f'{metric} - Modelo: {model}').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(algos) + 1)
    ws.row_dimensions[1].height = 22

    # Headers
    c = ws.cell(row=3, column=1, value='Dataset')
    c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER

    for j, algo in enumerate(algos, start=2):
        c = ws.cell(row=3, column=j, value=algo)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER

    # Datos (resaltando el mejor de cada fila)
    for i, dataset in enumerate(datasets, start=4):
        c = ws.cell(row=i, column=1, value=dataset)
        c.font, c.alignment, c.border = BOLD_FONT, LEFT, THIN_BORDER

        row_values = pivot.loc[dataset].values
        best_idx = int(row_values.argmax())

        for j, algo in enumerate(algos):
            val = float(pivot.loc[dataset, algo])
            c = ws.cell(row=i, column=j + 2, value=round(val, 4))
            c.font, c.alignment, c.border = CELL_FONT, CENTER, THIN_BORDER
            c.number_format = '0.0000'
            if j == best_idx:
                c.fill = BEST_FILL
                c.font = BOLD_FONT

    # Anchos y freeze
    ws.column_dimensions['A'].width = 22
    for j in range(2, len(algos) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.freeze_panes = 'B4'


def write_summary_sheet(wb: Workbook, df_long: pd.DataFrame, model: str) -> None:
    """Hoja resumen: el mejor algoritmo por dataset para cada métrica."""
    ws = wb.create_sheet('Resumen', 0)

    datasets = sorted(df_long['Dataset'].unique())
    metrics = ['ARI', 'AMI', 'NMI', 'Silhouette']

    ws.cell(row=1, column=1, value=f'Resumen - Mejor algoritmo por dataset ({model})').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    headers = ['Dataset'] + [f'Mejor en {m}' for m in metrics]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER

    for i, dataset in enumerate(datasets, start=4):
        c = ws.cell(row=i, column=1, value=dataset)
        c.font, c.border = BOLD_FONT, THIN_BORDER

        sub = df_long[df_long['Dataset'] == dataset]
        for j, m in enumerate(metrics, start=2):
            best_row = sub.loc[sub[m].idxmax()]
            text = f"{best_row['Algoritmo']} ({best_row[m]:.3f})"
            c = ws.cell(row=i, column=j, value=text)
            c.font, c.alignment, c.border = CELL_FONT, CENTER, THIN_BORDER

    ws.column_dimensions['A'].width = 22
    for j in range(2, 6):
        ws.column_dimensions[get_column_letter(j)].width = 22


def write_long_data_sheet(wb: Workbook, df_long: pd.DataFrame) -> None:
    """Hoja con todos los datos en formato largo (fácil de filtrar/ordenar)."""
    ws = wb.create_sheet('Datos_completos')

    for j, col in enumerate(df_long.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER

    for i, (_, row) in enumerate(df_long.iterrows(), start=2):
        for j, col in enumerate(df_long.columns, start=1):
            val = row[col]
            c = ws.cell(row=i, column=j, value=val)
            c.font = CELL_FONT
            c.alignment = LEFT if col == 'Dataset' else CENTER
            c.border = THIN_BORDER
            if isinstance(val, float):
                c.number_format = '0.0000'

    for j, col in enumerate(df_long.columns, start=1):
        width = max(len(col), df_long[col].astype(str).str.len().max()) + 2
        ws.column_dimensions[get_column_letter(j)].width = int(width)
    ws.freeze_panes = 'A2'


def build_single_model_excel(df_long: pd.DataFrame, model: str, out_path: Path) -> None:
    """Genera el Excel completo para un modelo."""
    wb = Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, df_long, model)
    for metric in ['ARI', 'AMI', 'NMI', 'Silhouette']:
        write_metric_sheet(wb, metric, df_long, model)
    write_long_data_sheet(wb, df_long)

    wb.save(out_path)
    print(f'✓ {out_path.name} guardado')


# ============================================================================
# Comparativa entre modelos
# ============================================================================

def build_all_models_excel(df_all: pd.DataFrame, out_path: Path) -> None:
    """
    Genera Excel comparando todos los modelos.
    Una hoja por métrica, con (Dataset, Algoritmo) en filas y modelos en columnas.
    """
    wb = Workbook()
    wb.remove(wb.active)

    models = sorted(df_all['Modelo'].unique())
    datasets = sorted(df_all['Dataset'].unique())
    algos = [a for a in ALGO_FILES if a in df_all['Algoritmo'].unique()]

    for metric in ['ARI', 'AMI', 'NMI', 'Silhouette']:
        ws = wb.create_sheet(metric)
        ws.cell(row=1, column=1, value=f'{metric} - Comparativa entre modelos').font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(models) + 2)

        # Headers
        c = ws.cell(row=3, column=1, value='Dataset')
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER
        c = ws.cell(row=3, column=2, value='Algoritmo')
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER
        for j, m in enumerate(models, start=3):
            c = ws.cell(row=3, column=j, value=m)
            c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN_BORDER

        # Datos
        row_i = 4
        for dataset in datasets:
            for algo in algos:
                sub = df_all[(df_all['Dataset'] == dataset) & (df_all['Algoritmo'] == algo)]
                if sub.empty:
                    continue
                ws.cell(row=row_i, column=1, value=dataset).border = THIN_BORDER
                ws.cell(row=row_i, column=2, value=algo).border = THIN_BORDER

                # Encontrar el mejor modelo para esta combinación
                vals_by_model = {m: float(sub[sub['Modelo'] == m][metric].iloc[0])
                                 for m in models if not sub[sub['Modelo'] == m].empty}
                if vals_by_model:
                    best_model = max(vals_by_model, key=vals_by_model.get)
                else:
                    best_model = None

                for j, m in enumerate(models, start=3):
                    val = vals_by_model.get(m)
                    if val is None:
                        continue
                    c = ws.cell(row=row_i, column=j, value=round(val, 4))
                    c.font, c.alignment, c.border = CELL_FONT, CENTER, THIN_BORDER
                    c.number_format = '0.0000'
                    if m == best_model:
                        c.fill = BEST_FILL
                        c.font = BOLD_FONT
                row_i += 1

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 14
        for j in range(3, len(models) + 3):
            ws.column_dimensions[get_column_letter(j)].width = 12
        ws.freeze_panes = 'C4'

    wb.save(out_path)
    print(f'✓ {out_path.name} guardado')


# ============================================================================
# Main
# ============================================================================

def discover_available_models() -> list[str]:
    """Detecta automáticamente los modelos disponibles bajo predictions/."""
    if not PREDICTIONS_DIR.exists():
        return []
    models = []
    for sub in PREDICTIONS_DIR.iterdir():
        if sub.is_dir() and any(sub.glob('pred_*.csv')):
            models.append(sub.name.upper())
    return sorted(models)


def main() -> None:
    parser = argparse.ArgumentParser(description='Construye tabla de resultados')
    parser.add_argument('--model', type=str, default=None,
                        help='Modelo a procesar (ej: CLIP, BLIP). '
                             'Si se omite, requiere --all.')
    parser.add_argument('--all', action='store_true',
                        help='Procesa todos los modelos detectados y genera '
                             'también la comparativa.')
    args = parser.parse_args()

    if not args.model and not args.all:
        parser.error('Especifica --model <NOMBRE> o --all')

    if args.all:
        available = discover_available_models()
        if not available:
            print('No se detectaron subdirectorios de modelos en '
                  f'{PREDICTIONS_DIR}. Crea predictions/clip/, predictions/blip/, etc.')
            sys.exit(1)
        print(f'Modelos detectados: {available}')

        # Cargar todos
        df_list = []
        for m in available:
            print(f'\nProcesando {m}...')
            df_m = load_model_results(m)
            df_list.append(df_m)
            df_m.to_csv(f'results_{m.lower()}_long.csv', index=False)
            print(f'  ✓ results_{m.lower()}_long.csv')
            build_single_model_excel(df_m, m, Path(f'results_{m.lower()}.xlsx'))

        # Comparativa
        df_all = pd.concat(df_list, ignore_index=True)
        df_all.to_csv('results_all_models_long.csv', index=False)
        print(f'\n✓ results_all_models_long.csv ({len(df_all)} filas)')
        build_all_models_excel(df_all, Path('results_all_models.xlsx'))
    else:
        df = load_model_results(args.model)
        df.to_csv(f'results_{args.model.lower()}_long.csv', index=False)
        print(f'✓ results_{args.model.lower()}_long.csv')
        build_single_model_excel(df, args.model, Path(f'results_{args.model.lower()}.xlsx'))


if __name__ == '__main__':
    main()
